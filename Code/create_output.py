# create_output.py - creates the output Excel file with the results

from pathlib import Path
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font

import os
import sys
import subprocess

import settings

colors = {
    "Strong Buy": settings.STRONG_BUY,
    "Buy": settings.BUY,
    "Hold": settings.HOLD,
    "Sell": settings.SELL,
    "Strong Sell": settings.STRONG_SELL
}

# output(rows, columns) creates the output Excel file within the Stock-Screener directory.
# O(n)
def output(rows, columns):
    BASE_DIR = Path(__file__).resolve().parent.parent  # Stock-Screener
    export_df = pd.DataFrame(rows, columns=columns)

    name = settings.FILE_NAME
    OUTPUT_FILE = BASE_DIR / f"{name} Output.xlsx"
    export_df.to_excel(OUTPUT_FILE, index=False)

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    TOP_COLOR = settings.TOP_COLOR                             # Header
    ROW_COLORS = (settings.ROW_1_COLOR, settings.ROW_2_COLOR)  # Alternating row colors

    # Color the header
    header_fill = PatternFill(start_color=TOP_COLOR, end_color=TOP_COLOR, fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill

    # Color the rating column (index 4) based on value
    rating_col_index = 4
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell = row[rating_col_index]
        rating = str(cell.value)
        if rating in colors:
            fill = PatternFill(start_color=colors[rating], end_color=colors[rating], fill_type="solid")
            cell.fill = fill
            cell.font = Font(color="FFFFFF")    # White text

    wb.save(OUTPUT_FILE)
    # Done coloring

    if settings.OPEN_ON_COMPLETION:
        if sys.platform.startswith("win"):      # Windows
            os.startfile(OUTPUT_FILE)
        elif sys.platform == "darwin":          # macOS
            subprocess.run(["open", OUTPUT_FILE])
        else:                                   # Linux
            subprocess.run(["xdg-open", OUTPUT_FILE])